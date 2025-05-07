# streamlit_biogas_balance.py
import streamlit as st
import math
import datetime
from io import BytesIO

# --- LIBRERÍAS DE EXPORTACIÓN ---
OPENPYXL_AVAILABLE = False
FPDF_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    pass

# --- FUNCIONES DE CÁLCULO (sin cambios) ---
def calcular_dimensiones_digestor(caudal_sustrato_kg_dia, trh_dias, densidad_sustrato_kg_m3=1000):
    volumen_sustrato_diario_m3 = caudal_sustrato_kg_dia / densidad_sustrato_kg_m3
    volumen_digestor_m3 = volumen_sustrato_diario_m3 * trh_dias
    diametro_digestor_m = altura_digestor_m = area_superficial_digestor_m2 = 0.0
    if volumen_digestor_m3 > 0:
        diametro_digestor_m = (4 * volumen_digestor_m3 / math.pi)**(1/3)
        altura_digestor_m = diametro_digestor_m
        area_superficial_digestor_m2 = 1.5 * math.pi * (diametro_digestor_m**2)
    return {
        "volumen_digestor_m3": volumen_digestor_m3,
        "diametro_digestor_m": diametro_digestor_m,
        "altura_digestor_m": altura_digestor_m,
        "area_superficial_digestor_m2": area_superficial_digestor_m2
    }

def realizar_calculos_balance(inputs_calc):
    results = {}
    caudal_sustrato_kg_dia = inputs_calc['caudal_sustrato_kg_dia']
    st_porcentaje = inputs_calc['st_porcentaje']
    sv_de_st_porcentaje = inputs_calc['sv_de_st_porcentaje']
    bmp_nm3_ch4_kg_sv = inputs_calc['bmp_nm3_ch4_kg_sv']
    eficiencia_digestion_porcentaje = inputs_calc['eficiencia_digestion_porcentaje']
    ch4_en_biogas_porcentaje = inputs_calc['ch4_en_biogas_porcentaje']
    cp_sustrato_kj_kg_c = inputs_calc['cp_sustrato_kj_kg_c']
    temp_op_digestor_c = inputs_calc['temp_op_digestor_c']
    temp_sustrato_entrada_c = inputs_calc['temp_sustrato_entrada_c']
    u_digestor_w_m2_k = inputs_calc['u_digestor_w_m2_k']
    area_superficial_digestor_m2 = inputs_calc['area_superficial_digestor_m2']
    temp_ambiente_promedio_c = inputs_calc['temp_ambiente_promedio_c']
    uso_biogas_opcion_idx = inputs_calc['uso_biogas_opcion_idx']
    chp_eficiencia_electrica_porcentaje = inputs_calc.get('chp_eficiencia_electrica_porcentaje', 0)
    chp_eficiencia_termica_porcentaje = inputs_calc.get('chp_eficiencia_termica_porcentaje', 0)
    caldera_eficiencia_porcentaje = inputs_calc.get('caldera_eficiencia_porcentaje', 0)
    consumo_electrico_aux_kwh_ton_sustrato = inputs_calc['consumo_electrico_aux_kwh_ton_sustrato']

    results['sv_alimentado_kg_dia'] = caudal_sustrato_kg_dia * (st_porcentaje / 100) * (sv_de_st_porcentaje / 100)
    results['ch4_producido_nm3_dia'] = results['sv_alimentado_kg_dia'] * bmp_nm3_ch4_kg_sv * (eficiencia_digestion_porcentaje / 100)
    results['biogas_producido_nm3_dia'] = 0
    if ch4_en_biogas_porcentaje > 0:
        results['biogas_producido_nm3_dia'] = results['ch4_producido_nm3_dia'] / (ch4_en_biogas_porcentaje / 100)
    pci_ch4_mj_nm3 = 35.8
    results['pci_biogas_mj_nm3'] = pci_ch4_mj_nm3 * (ch4_en_biogas_porcentaje / 100)
    results['energia_bruta_biogas_mj_dia'] = results['biogas_producido_nm3_dia'] * results['pci_biogas_mj_nm3']
    results['energia_bruta_biogas_kwh_dia'] = results['energia_bruta_biogas_mj_dia'] / 3.6
    results['calor_calentar_sustrato_mj_dia'] = (caudal_sustrato_kg_dia * cp_sustrato_kj_kg_c * (temp_op_digestor_c - temp_sustrato_entrada_c)) / 1000
    delta_t_digestor_ambiente = temp_op_digestor_c - temp_ambiente_promedio_c
    results['perdidas_calor_digestor_mj_dia'] = 0.0
    if delta_t_digestor_ambiente > 0 and area_superficial_digestor_m2 > 0:
        results['perdidas_calor_digestor_mj_dia'] = (u_digestor_w_m2_k * area_superficial_digestor_m2 * delta_t_digestor_ambiente * 3600 * 24) / 1000000
    results['demanda_termica_total_digestor_mj_dia'] = results['calor_calentar_sustrato_mj_dia'] + results['perdidas_calor_digestor_mj_dia']
    results['demanda_termica_total_digestor_kwh_dia'] = results['demanda_termica_total_digestor_mj_dia'] / 3.6
    results['electricidad_generada_bruta_kwh_dia'] = 0.0
    results['calor_util_generado_mj_dia'] = 0.0
    if uso_biogas_opcion_idx == 0: # CHP
        results['electricidad_generada_bruta_kwh_dia'] = results['energia_bruta_biogas_kwh_dia'] * (chp_eficiencia_electrica_porcentaje / 100)
        results['calor_util_generado_mj_dia'] = results['energia_bruta_biogas_mj_dia'] * (chp_eficiencia_termica_porcentaje / 100)
    elif uso_biogas_opcion_idx == 1: # Caldera
        results['calor_util_generado_mj_dia'] = results['energia_bruta_biogas_mj_dia'] * (caldera_eficiencia_porcentaje / 100)
    results['consumo_electrico_aux_total_kwh_dia'] = (caudal_sustrato_kg_dia / 1000) * consumo_electrico_aux_kwh_ton_sustrato
    results['electricidad_neta_exportable_kwh_dia'] = results['electricidad_generada_bruta_kwh_dia'] - results['consumo_electrico_aux_total_kwh_dia']
    results['calor_neto_disponible_mj_dia'] = results['calor_util_generado_mj_dia'] - results['demanda_termica_total_digestor_mj_dia']
    results['calor_neto_disponible_kwh_dia'] = results['calor_neto_disponible_mj_dia'] / 3.6
    return results

# --- INTERFAZ DE STREAMLIT ---
st.set_page_config(page_title="Balance Energético Biogás", layout="wide", page_icon="🔥") # Icono de fuego

st.title("🔥 Balance Energético Planta de Biogás") # Título y emoji actualizados
st.markdown("Esta aplicación realiza un balance de energía preliminar para una planta de biogás en fase de diseño.")
st.markdown("---")

# --- Parámetros de Configuración del Proyecto (en el área principal) ---
st.header("Parámetros de Configuración del Proyecto") # Título de la sección
col_proj_main1, col_proj_main2 = st.columns(2) # Usar 'main' para diferenciar de otras columnas
with col_proj_main1:
    project_name = st.text_input("Nombre del Proyecto", "Mi Planta de Biogás", key="project_name_main")
with col_proj_main2:
    analyst_name = st.text_input("Nombre del Analista", "Equipo de Diseño", key="analyst_name_main")
current_date = datetime.date.today().strftime("%Y-%m-%d")
st.caption(f"Fecha del análisis: {current_date}")
st.markdown("---")

# --- ENTRADAS DEL USUARIO EN LA BARRA LATERAL (EL RESTO DE PARÁMETROS) ---
st.sidebar.header("Parámetros de Entrada Detallados")

st.sidebar.subheader("1. Características del Sustrato")
sustrato_nombre_input = st.sidebar.text_input("Nombre/Tipo de sustrato", "Residuos Agroindustriales", key="sustrato_nombre_sidebar")
caudal_sustrato_kg_dia = st.sidebar.number_input("Caudal de sustrato (kg/día)", min_value=0.0, value=10000.0, step=100.0, format="%.2f")
st_porcentaje = st.sidebar.number_input("Sólidos Totales (ST) en sustrato (%)", min_value=0.0, max_value=100.0, value=20.0, step=0.1, format="%.1f")
sv_de_st_porcentaje = st.sidebar.number_input("Sólidos Volátiles (SV) como % de ST (%)", min_value=0.0, max_value=100.0, value=80.0, step=0.1, format="%.1f")
temp_sustrato_entrada_c = st.sidebar.number_input("Temperatura de entrada del sustrato (°C)", value=15.0, step=0.5, format="%.1f")
cp_sustrato_kj_kg_c = 4.186

bmp_fuente_opciones = ["Valor de laboratorio", "Estimación de literatura"]
bmp_fuente_seleccionada_texto = st.sidebar.selectbox("Fuente del BMP", bmp_fuente_opciones, help="Seleccione cómo se obtiene el Potencial Bioquímico de Metano.")
if "Valor de laboratorio" in bmp_fuente_seleccionada_texto:
    bmp_nm3_ch4_kg_sv = st.sidebar.number_input("BMP (Nm³ CH₄ / kg SV añadido)", min_value=0.0, value=0.35, step=0.01, format="%.2f")
else:
    bmp_nm3_ch4_kg_sv = st.sidebar.number_input("BMP estimado de literatura (Nm³ CH₄ / kg SV añadido)", min_value=0.0, value=0.30, step=0.01, format="%.2f")

st.sidebar.subheader("2. Diseño del Proceso de Digestión")
temp_op_digestor_opciones_dict = {"Mesofílico (~37-42 °C)": 38.0, "Termofílico (~50-55 °C)": 52.0}
temp_op_digestor_texto_sel = st.sidebar.selectbox("Rango de temperatura del digestor", list(temp_op_digestor_opciones_dict.keys()))
temp_op_digestor_c = temp_op_digestor_opciones_dict[temp_op_digestor_texto_sel]
st.sidebar.caption(f"Temperatura de operación seleccionada: {temp_op_digestor_c}°C")

eficiencia_digestion_porcentaje = st.sidebar.number_input("Eficiencia de digestión estimada (%)", min_value=0.0, max_value=100.0, value=75.0, step=0.5, format="%.1f")
trh_dias = st.sidebar.number_input("Tiempo de Retención Hidráulica (TRH) (días)", min_value=1.0, value=30.0, step=1.0, format="%.1f")
ch4_en_biogas_porcentaje = st.sidebar.number_input("Contenido de Metano (CH₄) estimado en biogás (%)", min_value=0.0, max_value=100.0, value=60.0, step=0.1, format="%.1f")

st.sidebar.markdown("###### Pérdidas Térmicas del Digestor")
temp_ambiente_promedio_c = st.sidebar.number_input("Temperatura ambiente promedio anual (°C)", value=10.0, step=0.5, format="%.1f")
u_digestor_w_m2_k = st.sidebar.number_input("Coef. global transf. calor (U) digestor (W/m²K)", min_value=0.0, value=0.5, step=0.01, format="%.2f", help="Ej: Aislado: 0.3-0.8; No aislado: 1.5-3.0")

st.sidebar.subheader("3. Utilización del Biogás")
uso_biogas_opciones_lista = ["Cogeneración (CHP)", "Caldera", "Upgrading a Biometano"]
uso_biogas_seleccionado_texto = st.sidebar.selectbox("Principal uso del biogás", uso_biogas_opciones_lista)
uso_biogas_opcion_idx = uso_biogas_opciones_lista.index(uso_biogas_seleccionado_texto)

chp_eficiencia_electrica_porcentaje = 0.0
chp_eficiencia_termica_porcentaje = 0.0
caldera_eficiencia_porcentaje = 0.0

if uso_biogas_opcion_idx == 0:
    chp_eficiencia_electrica_porcentaje = st.sidebar.number_input("Eficiencia eléctrica del CHP (%)", min_value=0.0, max_value=100.0, value=35.0, step=0.1, format="%.1f", key="chp_elec_eff")
    chp_eficiencia_termica_porcentaje = st.sidebar.number_input("Eficiencia térmica recuperable del CHP (%)", min_value=0.0, max_value=100.0, value=45.0, step=0.1, format="%.1f", key="chp_therm_eff")
elif uso_biogas_opcion_idx == 1:
    caldera_eficiencia_porcentaje = st.sidebar.number_input("Eficiencia de la caldera de biogás (%)", min_value=0.0, max_value=100.0, value=85.0, step=0.1, format="%.1f", key="boiler_eff")

st.sidebar.subheader("4. Consumos Energéticos Auxiliares")
consumo_electrico_aux_kwh_ton_sustrato = st.sidebar.number_input("Consumo eléctrico aux. (kWh / ton sustrato)", min_value=0.0, value=30.0, step=1.0, format="%.1f")

# --- Botón para ejecutar cálculos ---
st.markdown("---")
calcular_button = st.button("📊 RESULTADOS BALANCE ENERGÍA", type="primary", use_container_width=True)

if 'show_results' not in st.session_state:
    st.session_state.show_results = False

if calcular_button:
    st.session_state.show_results = True # Mostrar resultados cuando se presiona

if st.session_state.show_results:
    dim_digestor = calcular_dimensiones_digestor(caudal_sustrato_kg_dia, trh_dias)
    inputs_balance = {
        'sustrato_nombre': sustrato_nombre_input,
        'caudal_sustrato_kg_dia': caudal_sustrato_kg_dia,
        'st_porcentaje': st_porcentaje,
        'sv_de_st_porcentaje': sv_de_st_porcentaje,
        'bmp_nm3_ch4_kg_sv': bmp_nm3_ch4_kg_sv,
        'bmp_fuente_texto': bmp_fuente_seleccionada_texto,
        'eficiencia_digestion_porcentaje': eficiencia_digestion_porcentaje,
        'ch4_en_biogas_porcentaje': ch4_en_biogas_porcentaje,
        'cp_sustrato_kj_kg_c': cp_sustrato_kj_kg_c,
        'temp_op_digestor_c': temp_op_digestor_c,
        'temp_op_digestor_texto': temp_op_digestor_texto_sel,
        'temp_sustrato_entrada_c': temp_sustrato_entrada_c,
        'u_digestor_w_m2_k': u_digestor_w_m2_k,
        'area_superficial_digestor_m2': dim_digestor['area_superficial_digestor_m2'],
        'temp_ambiente_promedio_c': temp_ambiente_promedio_c,
        'uso_biogas_opcion_idx': uso_biogas_opcion_idx,
        'uso_biogas_texto': uso_biogas_seleccionado_texto,
        'chp_eficiencia_electrica_porcentaje': chp_eficiencia_electrica_porcentaje,
        'chp_eficiencia_termica_porcentaje': chp_eficiencia_termica_porcentaje,
        'caldera_eficiencia_porcentaje': caldera_eficiencia_porcentaje,
        'consumo_electrico_aux_kwh_ton_sustrato': consumo_electrico_aux_kwh_ton_sustrato,
        'trh_dias': trh_dias
    }
    results = realizar_calculos_balance(inputs_balance)

    st.header("Resultados del Balance") # Título de la sección de resultados
    st.markdown(f"Resultados para el proyecto: **{project_name}**")
    st.markdown("---")

    col_res1, col_res2, col_res3 = st.columns(3)
    with col_res1:
        st.subheader("Dimensiones del Digestor")
        st.metric("Volumen Estimado", f"{dim_digestor['volumen_digestor_m3']:.2f} m³")
        st.write(f"Diámetro Estimado (H=D): {dim_digestor['diametro_digestor_m']:.2f} m")
        st.write(f"Área Superficial Estimada: {dim_digestor['area_superficial_digestor_m2']:.2f} m²")
    with col_res2:
        st.subheader("Producción de Biogás")
        st.metric("Biogás Total Producido", f"{results.get('biogas_producido_nm3_dia', 0.0):.2f} Nm³/día")
        st.write(f"Metano (CH₄) producido: {results.get('ch4_producido_nm3_dia',0.0):.2f} Nm³/día")
        st.write(f"PCI del biogás: {results.get('pci_biogas_mj_nm3',0.0):.2f} MJ/Nm³")
        st.write(f"Energía Bruta en Biogás: {results.get('energia_bruta_biogas_mj_dia',0.0):.2f} MJ/día ({results.get('energia_bruta_biogas_kwh_dia',0.0):.2f} kWh/día)")
    with col_res3:
        st.subheader("Demanda Térmica del Digestor")
        st.metric("Demanda Térmica TOTAL", f"{results.get('demanda_termica_total_digestor_mj_dia',0.0):.2f} MJ/día", f"({results.get('demanda_termica_total_digestor_kwh_dia',0.0):.2f} kWh/día)")
        st.write(f"Calor para calentar sustrato: {results.get('calor_calentar_sustrato_mj_dia',0.0):.2f} MJ/día")
        st.write(f"Pérdidas de calor del digestor: {results.get('perdidas_calor_digestor_mj_dia',0.0):.2f} MJ/día")

    st.markdown("---")
    st.subheader("Producción y Consumos Energéticos")
    col_prod_res1, col_prod_res2 = st.columns(2)
    with col_prod_res1:
        st.write(f"**Uso Principal del Biogás:** {uso_biogas_seleccionado_texto}")
        if uso_biogas_opcion_idx == 0:
            st.metric("Electricidad Bruta Generada (CHP)", f"{results.get('electricidad_generada_bruta_kwh_dia',0.0):.2f} kWh/día")
            st.metric("Calor Útil Recuperado (CHP)", f"{results.get('calor_util_generado_mj_dia',0.0):.2f} MJ/día")
        elif uso_biogas_opcion_idx == 1:
            st.metric("Calor Útil Generado (Caldera)", f"{results.get('calor_util_generado_mj_dia',0.0):.2f} MJ/día")
        else:
            st.info("El biogás se destina a upgrading. Consumos y producción de biometano no detallados aquí.")
    with col_prod_res2:
        st.metric("Consumo Eléctrico Auxiliar Estimado", f"{results.get('consumo_electrico_aux_total_kwh_dia',0.0):.2f} kWh/día")

    st.markdown("---")
    st.subheader("BALANCE NETO DE ENERGÍA")
    col_neto_res1, col_neto_res2 = st.columns(2)
    with col_neto_res1:
        st.markdown("#### Balance Eléctrico")
        if uso_biogas_opcion_idx == 0:
            st.metric("ELECTRICIDAD NETA EXPORTABLE", f"{results.get('electricidad_neta_exportable_kwh_dia',0.0):.2f} kWh/día")
            if results.get('electricidad_neta_exportable_kwh_dia',0.0) < 0:
                st.error("¡ATENCIÓN! Déficit eléctrico.")
        else:
            st.metric("ELECTRICIDAD NETA (Consumo)", f"{-results.get('consumo_electrico_aux_total_kwh_dia',0.0):.2f} kWh/día")
    with col_neto_res2:
        st.markdown("#### Balance Térmico")
        st.metric("CALOR NETO DISPONIBLE/DÉFICIT", f"{results.get('calor_neto_disponible_mj_dia',0.0):.2f} MJ/día", f"{results.get('calor_neto_disponible_kwh_dia',0.0):.2f} kWh/día")
        if results.get('calor_neto_disponible_mj_dia',0.0) < 0:
            st.error(f"¡ATENCIÓN! Déficit térmico. Se necesitan {abs(results.get('calor_neto_disponible_mj_dia',0.0)):.2f} MJ/día adicionales.")
        elif results.get('calor_neto_disponible_mj_dia',0.0) > 0 and (uso_biogas_opcion_idx == 0 or uso_biogas_opcion_idx ==1):
            st.success("Calor excedentario disponible para otros usos.")

    st.sidebar.markdown("---")
    st.sidebar.header("Exportar Resultados")
    project_info_dict = {"nombre": project_name, "analista": analyst_name, "fecha": current_date}
    
    def sanitize_text_for_fpdf(text):
        """Reemplaza caracteres problemáticos para FPDF con la fuente Arial por defecto."""
        if not isinstance(text, str):
            text = str(text)
        # FPDF con fuentes estándar (latin-1) no maneja bien todos los caracteres Unicode.
        # Esta es una forma simple de reemplazar algunos comunes. Para soporte completo,
        # se necesitaría usar fuentes TrueType (TTF) que soporten Unicode.
        replacements = {
            '€': 'EUR', 'ñ': 'n', 'Ñ': 'N', 'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'ü': 'u', 'Ü': 'U', '¿': '?', '¡': '!'
            # Añade más reemplazos según sea necesario
        }
        for original, replacement in replacements.items():
            text = text.replace(original, replacement)
        # Intenta codificar a latin-1, reemplazando caracteres no mapeables
        return text.encode('latin-1', 'replace').decode('latin-1')

    def generar_excel_bytes(all_inputs, results_dict, dim_digestor_dict, project_info):
        # ... (código de generar_excel_bytes como antes, usando .get() para seguridad)
        if not OPENPYXL_AVAILABLE: 
            st.sidebar.warning("Exportación a Excel no disponible (falta 'openpyxl').")
            return None
        wb = Workbook()
        ws = wb.active
        # ... (resto de la función como antes, asegurándose de usar all_inputs.get(...) )
        # EJEMPLO de cómo usar .get():
        # add_excel_row(ws, ["Sustrato:", all_inputs.get('sustrato_nombre', 'N/A')])
        # Rellena el resto de la función de Excel como en la versión anterior,
        # asegurando que todas las claves de 'all_inputs' y 'results_dict' se accedan con .get()
        ws.title = "Resumen Balance Energético"
        header_font = Font(bold=True, size=12, color="00FFFFFF")
        category_font = Font(bold=True)
        bold_font = Font(bold=True)

        ws['A1'] = f"Balance Energético Preliminar: {project_info['nombre']}"
        ws.merge_cells('A1:D1'); ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Fecha: {project_info['fecha']}"; ws['A3'] = f"Analista: {project_info['analista']}"; ws.append([])
        
        def add_excel_row(sheet, data, font=None):
            sheet.append([sanitize_text_for_fpdf(str(d)) for d in data]) # Sanitizar para Excel también puede ser útil
            if font:
                for cell in sheet[sheet.max_row]: cell.font = font

        current_row_excel = ws.max_row + 1 
        add_excel_row(ws, ["PARÁMETROS DE ENTRADA"], font=header_font) 
        ws.merge_cells(start_row=current_row_excel, start_column=1, end_row=current_row_excel, end_column=3); current_row_excel +=1
        
        add_excel_row(ws, ["Sustrato:", all_inputs.get('sustrato_nombre', 'N/A')])
        add_excel_row(ws, ["Caudal Sustrato (kg/día):", all_inputs.get('caudal_sustrato_kg_dia', 0)])
        add_excel_row(ws, ["ST (%):", all_inputs.get('st_porcentaje',0)])
        # ... (muchos más add_excel_row)

        add_excel_row(ws, ["BALANCE NETO:"], font=category_font)
        add_excel_row(ws, ["  Electricidad Neta Exportable (kWh/día):", results_dict.get('electricidad_neta_exportable_kwh_dia',0)], font=bold_font)
        add_excel_row(ws, ["  Calor Neto Disponible/Déficit (MJ/día):", results_dict.get('calor_neto_disponible_mj_dia',0)], font=bold_font)

        for col_letter in ['A', 'B', 'C']: ws.column_dimensions[col_letter].width = 35 if col_letter == 'A' else 15
        
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream.getvalue()

    def generar_pdf_bytes(all_inputs, results_dict, dim_digestor_dict, project_info):
        if not FPDF_AVAILABLE:
            st.sidebar.warning("Exportación a PDF no disponible (falta 'fpdf2').")
            return None
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Intentar añadir una fuente Unicode si está disponible (requiere el archivo .ttf)
        # try:
        #     pdf.add_font("DejaVu", "", "DejaVuSansCondensed.ttf", uni=True)
        #     pdf.set_font("DejaVu", "", 10) # Usar la fuente Unicode
        # except RuntimeError:
        #     pdf.set_font("Arial", "", 10) # Fallback a Arial
        #     st.sidebar.caption("Nota PDF: Fuente DejaVu no encontrada, usando Arial (puede limitar caracteres).")
        
        pdf.set_font("Arial", "", 10) # Mantener Arial por simplicidad de dependencias

        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, sanitize_text_for_fpdf(f"Balance Energético Preliminar: {project_info['nombre']}"), 0, 1, "C")
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, sanitize_text_for_fpdf(f"Fecha: {project_info['fecha']} | Analista: {project_info['analista']}"), 0, 1, "C")
        pdf.ln(5)
        
        def add_pdf_section(title_pdf, data_dict_pdf):
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, sanitize_text_for_fpdf(title_pdf), 0, 1, "L")
            pdf.set_font("Arial", "", 9) # Tamaño de fuente más pequeño para contenido
            for key, value in data_dict_pdf.items():
                s_key = sanitize_text_for_fpdf(str(key))
                if isinstance(value, tuple):
                    s_val0 = sanitize_text_for_fpdf(str(value[0]))
                    s_val1 = sanitize_text_for_fpdf(str(value[1])) if len(value) > 1 else ''
                    line = f"  {s_key.ljust(45)}: {s_val0.ljust(15)} {s_val1}"
                else:
                    s_val = sanitize_text_for_fpdf(str(value))
                    line = f"  {s_key.ljust(45)}: {s_val}"
                
                # Limitar el ancho de multi_cell para evitar problemas
                # Ancho de página (210mm) - márgenes (15mm*2) = 180mm.
                # Damos un poco menos para seguridad.
                try:
                    pdf.multi_cell(170, 5, line, 0, "L") 
                except Exception as e_multicell:
                    print(f"Error en multi_cell con línea: '{line}'. Error: {e_multicell}")
                    pdf.multi_cell(170, 5, f"Error al renderizar: {s_key}", 0, "L") # Mostrar al menos la clave
            pdf.ln(3)

        input_data_pdf_content = {
            "Sustrato": all_inputs.get('sustrato_nombre', 'N/A'),
            "Caudal Sustrato (kg/día)": all_inputs.get('caudal_sustrato_kg_dia',0),
            "ST (%)": all_inputs.get('st_porcentaje',0),
            "SV (% de ST)": all_inputs.get('sv_de_st_porcentaje',0),
            "Fuente BMP": all_inputs.get('bmp_fuente_texto', 'N/A'),
            "BMP (Nm³ CH₄/kg SV)": all_inputs.get('bmp_nm3_ch4_kg_sv',0),
            "Temp. Op. Digestor (°C)": (all_inputs.get('temp_op_digestor_c',0), f"({all_inputs.get('temp_op_digestor_texto','N/A')})"),
            "Eficiencia Digestión (%)": all_inputs.get('eficiencia_digestion_porcentaje',0),
            "%CH₄ en biogás": all_inputs.get('ch4_en_biogas_porcentaje',0),
            "Uso Principal Biogás": all_inputs.get('uso_biogas_texto','N/A'),
        }
        if all_inputs.get('uso_biogas_opcion_idx') == 0:
            input_data_pdf_content["Eficiencia Eléctrica CHP (%)"] = all_inputs.get('chp_eficiencia_electrica_porcentaje',0)
            input_data_pdf_content["Eficiencia Térmica CHP (%)"] = all_inputs.get('chp_eficiencia_termica_porcentaje',0)
        elif all_inputs.get('uso_biogas_opcion_idx') == 1:
            input_data_pdf_content["Eficiencia Caldera (%)"] = all_inputs.get('caldera_eficiencia_porcentaje',0)
        add_pdf_section("PARÁMETROS DE ENTRADA", input_data_pdf_content)
        
        results_data_pdf_content = {
            "Dimensiones Digestor:": {
                "Volumen Estimado (m³)": f"{dim_digestor_dict.get('volumen_digestor_m3',0):.2f}",
                "Diámetro Estimado (m)": f"{dim_digestor_dict.get('diametro_digestor_m',0):.2f}",
                "Área Superficial (m²)": f"{dim_digestor_dict.get('area_superficial_digestor_m2',0):.2f}",
            },
            "Producción de Biogás:": {
                "Metano (CH₄) producido (Nm³/día)": f"{results_dict.get('ch4_producido_nm3_dia',0):.2f}",
                "Biogás total producido (Nm³/día)": f"{results_dict.get('biogas_producido_nm3_dia',0):.2f}",
                "Energía bruta en biogás (MJ/día)": f"{results_dict.get('energia_bruta_biogas_mj_dia',0):.2f}",
            },
            "Demanda Térmica Digestor:": {
                "Calor para calentar sustrato (MJ/día)": f"{results_dict.get('calor_calentar_sustrato_mj_dia',0):.2f}",
                "Pérdidas de calor del digestor (MJ/día)": f"{results_dict.get('perdidas_calor_digestor_mj_dia',0):.2f}",
                "Demanda térmica TOTAL (MJ/día)": f"{results_dict.get('demanda_termica_total_digestor_mj_dia',0):.2f}",
            },
            "Producción Energética (" + all_inputs.get('uso_biogas_texto','N/A') + "):": {
                "Electricidad bruta generada (kWh/día)": f"{results_dict.get('electricidad_generada_bruta_kwh_dia',0):.2f}" if all_inputs.get('uso_biogas_opcion_idx') == 0 else "N/A",
                "Calor útil generado (MJ/día)": f"{results_dict.get('calor_util_generado_mj_dia',0):.2f}",
            },
             "Consumos Auxiliares:":{
                "Consumo eléctrico auxiliar (kWh/día)": f"{results_dict.get('consumo_electrico_aux_total_kwh_dia',0):.2f}",
            },
            "BALANCE NETO:": {
                "ELECTRICIDAD NETA EXPORTABLE (kWh/día)": f"{results_dict.get('electricidad_neta_exportable_kwh_dia',0):.2f}" if all_inputs.get('uso_biogas_opcion_idx') == 0 else f"{-results_dict.get('consumo_electrico_aux_total_kwh_dia',0):.2f} (Consumo)",
                "CALOR NETO DISPONIBLE/DÉFICIT (MJ/día)": f"{results_dict.get('calor_neto_disponible_mj_dia',0):.2f}",
            }
        }
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, sanitize_text_for_fpdf("RESULTADOS DEL BALANCE (por día)"), 0, 1, "L")
        for section_title, data_items in results_data_pdf_content.items():
            pdf.set_font("Arial", "BU", 10)
            pdf.cell(0, 6, sanitize_text_for_fpdf(section_title), 0, 1, "L")
            pdf.set_font("Arial", "", 9)
            for key, value in data_items.items():
                s_key = sanitize_text_for_fpdf(str(key))
                s_val = sanitize_text_for_fpdf(str(value))
                line = f"  {s_key.ljust(50)}: {s_val}"
                try:
                    pdf.multi_cell(170, 5, line, 0, "L")
                except Exception as e_multicell_res:
                    print(f"Error en multi_cell resultados: '{line}'. Error: {e_multicell_res}")
                    pdf.multi_cell(170, 5, f"Error al renderizar: {s_key}", 0, "L")
            pdf.ln(2)

        pdf.ln(5); pdf.set_font("Arial", "B", 10); pdf.cell(0, 6, sanitize_text_for_fpdf("Notas Importantes:"), 0, 1, "L")
        pdf.set_font("Arial", "I", 9)
        pdf.multi_cell(170, 5, sanitize_text_for_fpdf(
            "- Este es un balance PRELIMINAR basado en estimaciones y supuestos.\n"
            "- Los valores de BMP, eficiencias y pérdidas pueden variar significativamente.\n"
            "- Se recomienda un análisis detallado con datos específicos del proyecto y de proveedores."
            ), 0, "L")
        
        try:
            return pdf.output(dest='S').encode('latin-1')
        except Exception as e_pdf_output:
            st.error(f"Error final al generar bytes del PDF: {e_pdf_output}")
            return None


    excel_data = generar_excel_bytes(inputs_balance, results, dim_digestor, project_info_dict)
    if excel_data:
        st.sidebar.download_button(
            label="📥 Descargar Resultados en Excel", data=excel_data,
            file_name=f"{project_name.replace(' ', '_')}_Balance_Energia_{current_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    pdf_data = generar_pdf_bytes(inputs_balance, results, dim_digestor, project_info_dict)
    if pdf_data:
        st.sidebar.download_button(
            label="📄 Descargar Resultados en PDF", data=pdf_data,
            file_name=f"{project_name.replace(' ', '_')}_Balance_Energia_{current_date}.pdf",
            mime="application/pdf"
        )
else:
    st.info("ℹ️ Configure los parámetros en la barra lateral y presione 'RESULTADOS BALANCE ENERGÍA' para ver el análisis.")

st.sidebar.markdown("---")
st.sidebar.info("Desarrollado como herramienta preliminar.")